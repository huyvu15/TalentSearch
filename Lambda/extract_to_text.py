import os
import textract
import openpyxl

# Define the file path
file_path = "a.pdf"     

# Get the file extension
file_extension = os.path.splitext(file_path)[1]

if file_extension in ['.xlsx', '.xls', '.csv']:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, values_only=True):
        print("".join(str(cell if cell is not None else "") for cell in row))
    workbook.close()
else:
    # Process the PDF or Word file
    text = textract.process(file_path)
    text = text.decode('utf-8')
    print(text)
    # with open("output.txt", "w", encoding="utf-8") as file:
    #     file.write(text)
